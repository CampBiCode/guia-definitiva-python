# Crear archivo de Excel
from openpyxl import Workbook

workbook = Workbook()
hoja = workbook.active # trabajar sobre la primer hoja

hoja["A1"] = "hola"
hoja["B1"] = "mundo"

workbook.save(filename = "hola_mundo.xlsx")



# Leer archivo de Excel
from openpyxl import load_workbook

workbook = load_workbook(filename = "muestra.xlsx")
print(workbook.sheetnames) # listado de nombres de las hojas del archivo

sheet = workbook.active # activar primer hoja
print(sheet)

print(sheet.title) # imprimir el nombre de la hoja

## acceder a valores de celdas
print(sheet["A1"].value)

## modo de lectura de archivos de Excel            
                                                    # abre archivo en modo lectura permitiendo leer archivos grandes
workbook = load_workbook(filename = "muestra.xlsx", read_only = True, data_only = True)
                                                                      # ignora todas las formulas usadas en los archivos y regresa solo sus valores

print(sheet["A1:C2"]) # acceder a un rango más grande

## iterar valores según posición
for row in sheet.iter_rows(min_row = 1,
                           max_row = 2,
                           min_col = 1,
                           max_col = 3):
    print([x.value for x in row])

### o
for valor in sheet.iter_rows(min_row = 1,
                           max_row = 2,
                           min_col = 1,
                           max_col = 3,
                           values_only = True):
    print(valor)



# Modificar archivos de Excel
## escribir en archivo nuevo
from openpyxl import Workbook

nombre_archivo = "datos.xlx"
sheet = workbook.active

sheet["A1"] = "Hola"
sheet["B1"] = "Mundo"

workbook.save(filename = nombre_archivo)

## operaciones básicas
### insertar columnas
                  # posición
sheet.insert_cols(idx=1, amount=1)
                         # cantidad

### eliminar columnas
sheet.delete_cols(idx=2, amount=5)

### insertar filas
sheet.insert_rows(idx = 1, amount = 3)

### eliminar filas
sheet.delete_rows(idx = 1, amount = 4) # eliminar 4 primeras filas

### cambiar hoja
from openpyxl import load_workbook

workbook = load_workbook(filename = "muestra.xlsx")
hoja_1 = workbook["Hoja1"]
hoja_2 = workbook["Hoja2"]

### renombrar hoja
hoja_1.title = "Hoja_uno"

### crear hoja
workbook.create_sheet("Nueva hoja")
workbook.sheetnames

### eliminar hoja
workbook.remove("Nueva hoja")
workbook.sheetnames



# Agregar formulas
from openpyxl import load_workbook

workbook = load_workbook(filename = "muestra.xlsx")
sheet = workbook.active
sheet["P2"] = "=AVERAGE(H2:H25)"
workbook.save(filename = "muestra.xlsx")